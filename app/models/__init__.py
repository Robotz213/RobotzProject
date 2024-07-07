from app import app
from app import db
import openpyxl
from datetime import datetime
from app.models.sql.sqlite3.auth import Users
from app.models.sql.mysql.law.processos import Processos

with app.app_context():

    db.create_all()
    root = Users.query.filter_by(login = '').first()
    
    if root is None:
        
        root_license = ""
        usuario = Users(
                    login = '',
                    nome_usuario = "",
                    senhacrip = "",
                    email = '',
                    type_user = "",
                    license_key = root_license)
        
        db.session.add(usuario)
    
    ws = openpyxl.load_workbook(r"R:\Github\RobotzProject\Sucessos 30-06-24.xlsx").active
    
    for i in range(2, ws.max_row):
        
        format_string = '%d-%m-%Y'
        numeroproc = ws.cell(row=i, column=1).value
        autor = ws.cell(row=i, column=2).value
        reu = ws.cell(row=i, column=3).value
        status = ws.cell(row=i, column=4).value
        distribuicao_data = ws.cell(row=i, column=5).value.replace("/", "-")
        distribuicao_data = datetime.strptime(distribuicao_data, format_string)
        valor_causa = ws.cell(row=i, column=6).value
        proc = Processos.query.filter(Processos.processo == numeroproc).first()
        
        if proc is None:
            
            processo = Processos(
                processo = numeroproc,
                autor = autor,
                reu = reu,
                status = status,
                distribuicao_data = distribuicao_data,
                valor_causa = valor_causa  
            )
            
            db.session.add(processo)
        
        if i == ws.max_row:
            break
    
    
        
    db.session.commit()